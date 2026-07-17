# prisma/process-taco.py
# Processa o Excel da TACO e gera taco_mapeada.json
# Execução: python3 prisma/process-taco.py
# Requer: pip install openpyxl

from openpyxl import load_workbook
import json, os

XLSX_PATH = os.path.join(os.path.dirname(__file__), 'Taco-4a-Edicao.xlsx')
OUT_PATH  = os.path.join(os.path.dirname(__file__), 'taco_mapeada.json')

VALID_CATS = {
    'Cereais e derivados',
    'Verduras, hortaliças e derivados',
    'Frutas e derivados',
    'Alimentos preparados',
    'Leguminosas e derivados',
    'Leite e derivados',
    'Carnes e derivados',
    'Gorduras e óleos',
    'Bebidas (alcoólicas e não alcoólicas)',
    'Nozes e sementes',
    'Miscelâneas',
    'Ovos e derivados',
    'Pescados e frutos do mar',
    'Produtos açucarados',
    'Outros alimentos industrializados',
}

# Alimentos marcados como favoritos (aparecem na aba padrão do modal)
FAVORITES = {
    'Arroz, integral, cozido', 'Arroz, tipo 1, cozido', 'Arroz, tipo 2, cozido',
    'Aveia, flocos, crua', 'Macarrão, espaguete, cozido',
    'Pão, de forma, tradicional', 'Pão francês', 'Pão, de forma, integral',
    'Batata-doce, cozida', 'Batata, inglesa, cozida', 'Mandioca, mansa, cozida',
    'Inhame, cozido', 'Mandioquinha, cozida',
    'Feijão, carioca, cozido', 'Feijão, preto, cozido',
    'Feijão, fradinho, cozido', 'Feijão, rajado, cozido',
    'Lentilha, cozida', 'Grão-de-bico, cozido', 'Ervilha, enlatada, drenada',
    'Abóbora, cabotian, cozida', 'Abobrinha, italiana, cozida', 'Abobrinha, italiana, crua',
    'Agrião, cru', 'Alface, crespa, crua', 'Alface, americana, crua',
    'Beterraba, cozida', 'Brócolis, cozido', 'Cenoura, crua', 'Cenoura, cozida',
    'Chuchu, cozido', 'Couve, manteiga, crua', 'Couve-flor, cozida',
    'Espinafre, cozido', 'Pepino, crua', 'Rúcula, crua', 'Tomate, cru',
    'Berinjela, cozida', 'Pimentão, verde, cru', 'Vagem, cozida', 'Repolho, cru',
    'Abacate, cru', 'Abacaxi, cru', 'Açaí, polpa, congelada',
    'Banana, nanica', 'Banana, prata',
    'Laranja, lima, crua', 'Laranja, pera, crua',
    'Maçã, fuji, crua', 'Mamão, formosa, cru', 'Mamão, havaí, cru',
    'Manga, tommy, crua', 'Melancia, crua', 'Melão, crua',
    'Morango, cru', 'Pera, crua', 'Uva, itália, crua', 'Goiaba, vermelha, crua',
    'Frango, peito, sem pele, assado', 'Frango, peito, sem pele, grelhado',
    'Frango, coxa, sem pele, assada', 'Frango, sobrecoxa, sem pele, assada',
    'Carne, bovina, patinho, sem gordura, grelhado',
    'Carne, bovina, alcatra, sem gordura, grelhada',
    'Carne, bovina, filé mignon, sem gordura, grelhado',
    'Carne, bovina, coxão mole, sem gordura, grelhado',
    'Carne, bovina, acém, moído, cozido',
    'Ovo, de galinha, inteiro, cozido/10minutos',
    'Ovo, de galinha, inteiro, frito',
    'Ovo, de galinha, clara, cozida/10minutos',
    'Atum, conserva em óleo',
    'Camarão, Rio Grande, grande, cozido',
    'Tilápia, filé, assado', 'Tilápia, filé, cozido',
    'Salmão, filé, assado', 'Salmão, filé, grelhado',
    'Sardinha, conserva em óleo', 'Sardinha, fresca, assada',
    'Iogurte, natural, desnatado', 'Iogurte, natural',
    'Leite, de vaca, desnatado', 'Leite, de vaca, integral',
    'Queijo, minas, frescal', 'Queijo, mozarela', 'Queijo, ricota',
    'Azeite, de oliva, extra virgem', 'Manteiga, com sal',
    'Óleo, de soja', 'Óleo, de canola',
    'Castanha-do-Brasil, crua', 'Castanha-de-caju, torrada, salgada',
    'Amendoim, torrado, salgado', 'Noz, crua',
    'Gergelim, semente', 'Linhaça, semente',
}

wb   = load_workbook(XLSX_PATH, read_only=True)
ws   = wb.active
foods = []
current_cat = None
taco_id = 0

for row in ws.iter_rows(min_row=4, values_only=True):
    col_a, col_b, _, col_d, _, col_f, col_g, _, col_i, col_j = row[:10]

    if isinstance(col_a, str) and col_a.strip() in VALID_CATS:
        current_cat = col_a.strip()
        continue

    if not col_b or not isinstance(col_b, str):
        continue
    if any(kw in col_b.lower() for kw in ['descrição', 'alimento', 'NA -', 'Tr -', '* -']):
        continue

    def sf(v):
        if v is None or v in ('NA','Tr','') or (isinstance(v, str) and (v.startswith('=') or not v.strip())):
            return 0.0
        try: return round(float(v), 1)
        except: return 0.0

    try:
        kcal = float(col_d) if col_d not in (None,'NA','Tr') else None
        if kcal is None: continue
    except: continue

    taco_id += 1
    name = col_b.strip()
    foods.append({
        'id':         taco_id,
        'name':       name,
        'category':   current_cat or 'Outros',
        'kcal':       sf(col_d),
        'p':          sf(col_f),
        'f':          sf(col_g),
        'c':          sf(col_i),
        'fiber':      sf(col_j),
        'isFavorite': name in FAVORITES,
    })

with open(OUT_PATH, 'w', encoding='utf-8') as f:
    json.dump(foods, f, ensure_ascii=False, indent=2)

favs = sum(1 for f in foods if f['isFavorite'])
print(f'✅ {len(foods)} alimentos processados → taco_mapeada.json')
print(f'   Favoritos: {favs} | TACO completa: {len(foods) - favs}')